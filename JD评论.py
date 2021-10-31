import requests
import json
import openpyxl

url = 'https://club.jd.com/comment/productPageComments.action?'
count = 1
params = {
    'callback': 'fetchJSON_comment98',
    'productId': '100008348530',
    'score': '0',
    'sortType': '5',
    'page': '0',
    'pageSize': '10',
    'isShadowSku': '0',
    'fold': '1'
}

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Safari/537.36'
}


def get_resp_json(productId, page):
    params['productId'] = productId
    params['page'] = page
    resp = requests.get(url=url, headers=headers, params=params)
    resp.encoding = 'gbk'
    return json.loads(resp.text.replace("fetchJSON_comment98(", "").replace(");", ""))


def get_comment(resp_dic):
    global count
    comments = resp_dic['comments']
    for com in comments:
        customer_id = com['id']
        comment_detail = com['content']
        creationTime = com['creationTime']
        count += 1
        outws.cell(row=count, column=1, value=str(count - 1))
        outws.cell(row=count, column=2, value=str(customer_id))
        outws.cell(row=count, column=3, value=str(comment_detail))
        outws.cell(row=count, column=4, value=creationTime)


if __name__ == '__main__':
    # 创建excel
    outwb = openpyxl.Workbook()
    outws = outwb.create_sheet(index=0)

    outws.cell(row=1, column=1, value="index")
    outws.cell(row=1, column=2, value="customer_id")
    outws.cell(row=1, column=3, value="comment")
    outws.cell(row=1, column=4, value="creationTime")
    resp_dic = []

    for i in range(1, 100, 2):
        # 100009077475 一样的
        resp = get_resp_json(productId=100009077475, page=i)
        resp_dic.append(resp)
        get_comment(resp)
        print(f"第{i}页评论数据保存")

    with open('JDiphone.json', mode='w', encoding='utf-8') as f:
        f.write(json.dumps(resp_dic, indent=2, ensure_ascii=False))

    outwb.save("iphone-评论.xlsx")  # 保存
