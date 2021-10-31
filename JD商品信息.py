import requests
from lxml import etree
import json
import openpyxl

count = 1
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36'
}


# 根据商品id获取评论数
def comment_count(prod_id):
    # https://club.jd.com/comment/productCommentSummaries.action?referenceIds=10028108187640&callback=jQuery8256910&_=1632991290851
    url = "https://club.jd.com/comment/productCommentSummaries.action?referenceIds=" + str(
        prod_id) + "&callback=jQuery8827474&_=1615298058081"
    res = requests.get(url, headers=headers)
    res.encoding = 'gbk'
    # print(res.text)
    text = (res.text).replace("jQuery8827474(", "").replace(");", "")
    # print(text)
    text = json.loads(text)
    comment_count = text['CommentsCount'][0]['CommentCountStr']

    comment_count = comment_count.replace("+", "")
    # 对“万”进行操作
    if "万" in comment_count:
        comment_count = comment_count.replace("万", "")
        comment_count = str(int(comment_count) * 10000)

    return comment_count


def get_page(page_all):
    page_t = 1
    s = 1
    for i in range(1, page_all):
        data['page'] = str(i)
        print("page=" + str(page_t) + ",s=" + str(s))
        get_list(url, data)
        page_t = page_t + 2
        s = s + 60


def get_list(url, data):
    global count
    res = requests.get(url=url, headers=headers, params=data)
    res.encoding = 'utf-8'
    text = res.text

    selector = etree.HTML(text)
    list = selector.xpath('//*[@id="J_goodsList"]/ul/li')
    # 获得商品信息
    for i in list:
        title = i.xpath('.//div[@class="p-name p-name-type-2"]/a/em/text()')[0]
        price = i.xpath('.//div[@class="p-price"]/strong/i/text()')[0]
        product_id = i.xpath('.//div[@class="p-commit"]/strong/a/@id')[0].replace("J_comment_", "")
        temp = [title, price, product_id]
        prod_list.append(temp)
        # 获取商品评论数
        commentcount = comment_count(product_id)
        print("title" + str(title))
        print("price=" + str(price))
        print("comment_count=" + commentcount)
        print("------------------------------------------------------")

        # print("title" + str(title))
        # print("price=" + str(price))
        # print("product_id=" + str(product_id))
        # print("------------------------------------------------------")
        count += 1
        outws.cell(row=count, column=1, value=str(count - 1))
        outws.cell(row=count, column=2, value=str(title))
        outws.cell(row=count, column=3, value=str(price))
        outws.cell(row=count, column=4, value=commentcount)


if __name__ == '__main__':
    data = {
        'keyword': '苹果',
        'wq': '苹果',
        'ev': '',
        'pvid': '5b11c104ebcb4fdeac462deaf66a2222',
        'page': '1',
        's': '1',
        'click': '0'
    }

    prod_list = []
    # url = "https://search.jd.com/search?keyword=笔记本&wq=笔记本&ev=exbrand_联想%5E&pvid=5b11c104ebcb4fdeac462deaf66a2222&page=2&s=1&click=0"
    url = "https://search.jd.com/search"
    outwb = openpyxl.Workbook()
    outws = outwb.create_sheet(index=0)

    outws.cell(row=1, column=1, value="index")
    outws.cell(row=1, column=2, value="title")
    outws.cell(row=1, column=3, value="price")
    outws.cell(row=1, column=4, value="CommentCount")
    # res = requests.get(url=url, headers=headers, params=data)
    # res.encoding = 'utf-8'
    # text = res.text
    #
    # selector = etree.HTML(text)
    # list = selector.xpath('//*[@id="J_goodsList"]/ul/li')
    #
    #
    # # 获得商品信息
    # for i in list:
    #     title = i.xpath('.//div[@class="p-name p-name-type-2"]/a/em/text()')[0]
    #     price = i.xpath('.//div[@class="p-price"]/strong/i/text()')[0]
    #     product_id = i.xpath('.//div[@class="p-commit"]/strong/a/@id')[0].replace("J_comment_", "")
    #     temp = [title, price, product_id]
    #     prod_list.append(temp)
    #     ###获取商品评论数
    #     commentcount = comment_count(product_id)
    #     print("title" + str(title))
    #     print("price=" + str(price))
    #     print("comment_count=" + commentcount)
    #     print("------------------------------------------------------")
    #
    #     # print("title" + str(title))
    #     # print("price=" + str(price))
    #     # print("product_id=" + str(product_id))
    #     # print("------------------------------------------------------")
    #     count += 1
    #     outws.cell(row=count, column=1, value=str(count - 1))
    #     outws.cell(row=count, column=2, value=str(title))
    #     outws.cell(row=count, column=3, value=str(price))
    #     outws.cell(row=count, column=4, value=commentcount)
    get_page(3)
    outwb.save("京东商品-苹果.xlsx")  # 保存
