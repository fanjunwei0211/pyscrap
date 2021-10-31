import requests
import openpyxl
import json
from time import sleep

# 一定要自己手动更改cookies，过一段时间没访问网站会自动失效

count = 1

comment_url = 'https://rate.tmall.com/list_detail_rate.htm?'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Safari/537.36',
    'referer': ''
}

cookies = {
    "Cookies": "lid=%E6%B2%90%E5%B9%B4%E7%81%AC%E6%9C%A8%E5%A2%93; enc=0JGU5%2FKpkQG9H6Z15veZ%2Bu0oVqpcAQc8hkwAG9t6K%2FIjq1p2svJ22S6afRQ%2F5mrs4Jz3r3Ukl02HT2QvmY1FxQ%3D%3D; cna=+lG/GZ6bGmQCAd7Rmz3Cq3vL; _m_h5_tk=2010134a863f0f82ce18e091db35f08f_1633024572965; _m_h5_tk_enc=bb8e61f07def631da54e0ecfe92eadf8; dnk=%5Cu6C90%5Cu5E74%5Cu706C%5Cu6728%5Cu5893; uc1=pas=0&cookie15=UtASsssmOIJ0bQ%3D%3D&existShop=false&cookie14=Uoe3dP2ZQt10Jw%3D%3D&cookie16=VFC%2FuZ9az08KUQ56dCrZDlbNdA%3D%3D&cookie21=WqG3DMC9Fb5mPLIQo9kR&cart_m=0; uc3=lg2=VFC%2FuZ9ayeYq2g%3D%3D&vt3=F8dCujaNo3CzyZg6ejg%3D&nk2=gIdYVPyZFz1gqA%3D%3D&id2=VyyUw9BPjgHPTQ%3D%3D; tracknick=%5Cu6C90%5Cu5E74%5Cu706C%5Cu6728%5Cu5893; uc4=nk4=0%40gsxc7D7%2B%2FIsZlBPFX4fIa1apcX4L&id4=0%40VXtbaq2Ipq%2Bq0lNh43gNKrahCQPe; lgc=%5Cu6C90%5Cu5E74%5Cu706C%5Cu6728%5Cu5893; login=true; cookie2=18e3e357ea639906c02764ae2cfea1d7; sgcookie=E100p85K4IhaFDxg7M0PPmFCqWx70ddRhDxQWDQnHnI4FhobfFbDkMDRPVlFubjagfyrnjsaVu5JaMDkWBzlR6oFzYk%2FAfIyMrMSng8z3cIvyzQ%3D; cancelledSubSites=empty; t=0d3d9b2589db76201478bad573278804; csg=5423233a; _tb_token_=e977531e83137; x5sec=7b22617365727665723b32223a2261653133363837313836616531373165353039366232383734376663663030304349536532346f4745507239373537716b74666159426f4d4e4441314f4455774d6a63354e7a73794d4b366838593043227d; isg=BMLCq7jEF8tnEzWXb5aBibhGE8gkk8atMnX2HgzUbjf3X27Z9SNYvI_ZD1sjDz5F; l=eBS8Nk8VOhG0qEmjBO5Zlurza77OopdXGsPzaNbMiIncC6mhp-JiZE-Qc4V2ACKRR8XcGZ8p44v7xneOlV-7Ji70xZyCwTvQbgHyQef.."
}

params = {'itemId': '601351169591',
          'spuId': '1322791524',
          'sellerId': '2206378320010',
          'order': '3',
          'currentPage': '2',
          'append': '0',
          'content': '1',
          'ua': '098#E1hv+QvXvJpvUvCkvvvvvjiWPsMvzj1bPFzvsjEUPmPWgj1WPLMpsj3EnLsptj3bdvhvmpmvGNxBvvmdwOvCvCcYcbxaCKMwznQ9nHdAZ0eCkRSRsR8+vpvEvUVQrm6vvU/RRvhvCvvvphmVvpvhvUCvpv9CvhAmDTA+jX7rV8TrEcqha4g7WdEYVVzwdiG4jcxX0f06W3vOJ1kHsfUpeC6AxYjxRLwp+3+daNpArqVTbZkt640AdX9apbmxdXkK5eUHuvhvmvvvpLz+hbSikvhvC9hvpyPy6b9Cvm9vvhh5MMMYNQvvBZZvvvVsvvCHBpvvv79vvhxHvvvC4vvvBZhvvvHuvvhvC9mvphvvvvgCvvpvvPMMRvhvCvvvphmevpvhvvCCBUvCvCLwMRaafrMwznQw+HS5vSQ8zv1444QCvvyvvoUeP9vv91ggvpvhvvCvp8QCvCT8vxXoSvvvurQtUx7x6fmsVtr1b+LhVCkwejCCsUe+vpvEvU2STgZvvmr3dvhvjIuvzn3gvvv+gFy3xXcB+bWfezhK1WpOVCeOwhcp34QCvvyv9nsOc9vvdCR+vpvo3vEiXT6vv2EJ+oY2rEZGRFeQdvhvhIpm8mfwvvvTGfhSpqoxP2It+8QCvC8hvyxbf9vvKIT1b6R6DLQTRfmvvpvZz215fqcNznswdWafzsSGjYAv7IhgvpvhvvCvpv==',
          'needFold': '0',
          '_ksTS': '1633077755553_782',
          'callback': 'jsonp783'}


def get_resp_json(url, Id, page):
    start_url = url.format(Id)
    headers['referer'] = start_url
    params['currentPage'] = page
    resp = requests.get(url=comment_url, headers=headers, cookies=cookies, params=params)
    resp.encoding = 'utf-8'
    # return resp.text
    return json.loads(resp.text.replace("jsonp783(", "").replace(")", ""))


def get_comment(resp_dic):
    global count
    comments = resp_dic['rateDetail']['rateList']
    for com in comments:
        customer_id = com['sellerId']
        comment_detail = com['rateContent']
        creationTime = com['rateDate']
        count += 1
        outws.cell(row=count, column=1, value=str(count - 1))
        outws.cell(row=count, column=2, value=str(customer_id))
        outws.cell(row=count, column=3, value=str(comment_detail))
        outws.cell(row=count, column=4, value=creationTime)


if __name__ == '__main__':
    all_dic = []
    # # 创建excel
    outwb = openpyxl.Workbook()
    outws = outwb.create_sheet(index=0)

    outws.cell(row=1, column=1, value="index")
    outws.cell(row=1, column=2, value="customer_id")
    outws.cell(row=1, column=3, value="comment")
    outws.cell(row=1, column=4, value="creationTime")

    url = "https://detail.tmall.com/item.htm?id={}&spm=a21bo.21814703.201876.1.624811d9xDS6TX&scm=1007.34127.227518.0&pvid=a60e0f1e-10c9-49ea-94a3-fec1f1c6c5c0"

    for i in range(1, 20):
        resp_json = get_resp_json(url, 601351169591, i)
        all_dic.append(resp_json)
        get_comment(resp_json)
        # print(resp_json)
        print(f"第{i}页获取成功！")
        sleep(4)

    with open('淘宝评论.json', mode='w', encoding='utf-8') as f:
        f.write(json.dumps(all_dic, indent=2, ensure_ascii=False))

    outwb.save("tb-评论.xlsx")  # 保存
